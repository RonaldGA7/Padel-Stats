import os
import re
import urllib.parse
import tempfile
import zipfile
from datetime import datetime, date
from typing import List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ==========================
# Configuración
# ==========================
SAQUE_ESTADOS = ["Correcto", "Primer error", "Doble falta"]
RESULTADOS = ["Winner", "Error forzado", "Error no forzado"]

GOLPES = [
    "Saque",
    "Smash", "Bandeja", "Víbora", "Globo",
    "Volea derecha", "Volea revés",
    "Derecha", "Revés", "Devolución",
    "Bajada pared derecha", "Bajada pared revés",
    "Salida pared derecha", "Salida pared revés",
    "Otro",
]

PTS_TEXT = {0: "0", 1: "15", 2: "30", 3: "40"}

SHEET_EVENTS = "Eventos"
SHEET_SUMMARY = "Resumen"

# ✅ (4) Sin TB_Eq1 / TB_Eq2
# ✅ (5) Sin WinnerDeSaque
EVENT_COLS = [
    "FechaHora",
    "Set",
    "Games_Eq1",
    "Games_Eq2",
    "Pts_Eq1",
    "Pts_Eq2",
    "TieBreak",
    "TB_Tipo",      # "SET" / "SUPER" / ""
    "Saca",
    "SaqueEstado",
    "Resultado",
    "Golpe",
    "JugadorActor",
    "JugadorProvocador",
    "Asistencia",
    "Asistente",
    "EquipoGanadorPunto",
    "DuracionPunto",
    "GoldenReceiver",
]


# ==========================
# CSS
# ==========================
def inject_css():
    st.markdown("""
<style>
div[data-testid="stSegmentedControl"] button[aria-pressed="true"]{
  background: #2E7D32 !important;
  color: white !important;
  border-color: #2E7D32 !important;
}
div[data-testid="stSegmentedControl"] button{
  padding: 0.30rem 0.55rem !important;
  border-radius: 10px !important;
  font-size: 0.90rem !important;
  line-height: 1.1rem !important;
  min-height: 2.0rem !important;
}
</style>
""", unsafe_allow_html=True)


def segmented_toggle(
    container,
    label: str,
    state_key: str,
    options: List[str],
    key: str,
    allow_clear: bool = True,
    disabled: bool = False,
) -> Tuple[str, str]:
    """
    Segmented con estado en session_state[state_key].
    IMPORTANTE: si allow_clear=True, aparece '—' y NO se auto-selecciona nada.
    """
    old = st.session_state.get(state_key, "")
    opts = (["—"] + options) if allow_clear else options

    if allow_clear:
        default_value = "—" if old == "" else (old if old in options else "—")
    else:
        default_value = old if old in options else options[0]

    val = container.segmented_control(
        label,
        options=opts,
        default=default_value,
        key=key,
        disabled=disabled,
    )

    new = "" if (allow_clear and val == "—") else val
    st.session_state[state_key] = new
    return new, old


# ==========================
# Helpers generales
# ==========================
def sanitize_filename(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[\\/:*?\"<>|]", "_", s)
    s = re.sub(r"\s+", " ", s)
    return s


def mmss_from_start(start_dt: Optional[datetime]) -> str:
    if start_dt is None:
        return "00:00"
    sec = int((datetime.now() - start_dt).total_seconds())
    sec = max(sec, 0)
    mm, ss = divmod(sec, 60)
    return f"{mm:02d}:{ss:02d}"


def unique_players(eq1: List[str], eq2: List[str]) -> List[str]:
    seen = set()
    out = []
    for x in eq1 + eq2:
        x = (x or "").strip()
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def equipo_de(jugador: str, eq1: List[str], eq2: List[str]) -> int:
    if jugador in eq1:
        return 1
    if jugador in eq2:
        return 2
    return 0


def opuesto(eq: int) -> int:
    return 2 if eq == 1 else 1 if eq == 2 else 0


def companero(actor: str, eq1: List[str], eq2: List[str]) -> str:
    if actor in eq1:
        return eq1[1] if eq1[0] == actor else eq1[0]
    if actor in eq2:
        return eq2[1] if eq2[0] == actor else eq2[0]
    return ""


def ganador_equipo_por_regla(resultado: str, actor: str, provocador: str, eq1: List[str], eq2: List[str]) -> int:
    if resultado == "Winner":
        return equipo_de(actor, eq1, eq2)
    if resultado == "Error no forzado":
        return opuesto(equipo_de(actor, eq1, eq2))
    if resultado == "Error forzado":
        return equipo_de(provocador, eq1, eq2)
    return 0


# ==========================
# Compartir (Guía móvil/tablet)
# ==========================
def normalizar_whatsapp(numero: str) -> str:
    return re.sub(r"\D", "", numero or "")


def mailto_link(subject: str, body: str, to_email: str = "") -> str:
    subject_q = urllib.parse.quote(subject or "")
    body_q = urllib.parse.quote(body or "")
    to_q = urllib.parse.quote(to_email or "")
    return f"mailto:{to_q}?subject={subject_q}&body={body_q}"


def whatsapp_link(phone_digits: str, message: str) -> str:
    msg_q = urllib.parse.quote(message or "")
    if phone_digits:
        return f"https://wa.me/{phone_digits}?text={msg_q}"
    return f"https://wa.me/?text={msg_q}"


def ui_compartir_excel_con_guia(excel_file: str):
    if not os.path.exists(excel_file):
        st.info("Todavía no hay Excel para compartir.")
        return

    st.subheader("📤 Compartir Excel (móvil/tablet)")
    st.caption(
        "Por seguridad, **Email y WhatsApp NO permiten adjuntar archivos automáticamente desde un link**. "
        "Flujo correcto: **descargar primero** y luego **adjuntar manualmente**."
    )

    st.markdown("### ✅ Paso 1: Descarga el Excel")
    with open(excel_file, "rb") as f:
        st.download_button(
            "⬇️ Descargar Excel",
            data=f,
            file_name=os.path.basename(excel_file),
            use_container_width=True,
            key="dl_excel_share",
        )

    st.info(
        "📱 **Dónde queda el archivo descargado**\n\n"
        "- **iPhone/iPad:** App **Archivos** → *Descargas* (o iCloud Drive/En mi iPhone según el navegador)\n"
        "- **Android:** **Files/Archivos** → *Downloads/Descargas*"
    )

    st.markdown("### ✅ Paso 2: Elige cómo enviarlo")
    asunto = "Resumen Partido Padel"
    mensaje_default = "Te comparto el resumen del partido. Adjunté el Excel."

    tab_mail, tab_wa = st.tabs(["📧 Email", "💬 WhatsApp"])

    with tab_mail:
        email_to = st.text_input("Email destinatario", placeholder="ej: alguien@mail.com", key="share_email_to")
        cuerpo = st.text_area("Mensaje", value=mensaje_default, height=120, key="share_email_body")
        link = mailto_link(asunto, cuerpo, email_to.strip())
        st.link_button("📧 Abrir Email", link, use_container_width=True)

        st.markdown("### ✅ Paso 3: Adjunta el Excel en la app de correo")
        st.markdown(
            "- **iPhone/iPad (Mail):** botón **Adjuntar** → **Archivos** → busca el Excel.\n"
            "- **Android (Gmail/Correo):** clip **📎** → **Adjuntar archivo** → **Descargas**."
        )

    with tab_wa:
        phone = st.text_input(
            "Número WhatsApp (con prefijo país)",
            placeholder="ej: +34 600 111 222",
            key="share_wa_phone",
        )
        wa_msg = st.text_area("Mensaje", value=mensaje_default, height=120, key="share_wa_msg")
        phone_digits = normalizar_whatsapp(phone)
        link = whatsapp_link(phone_digits, wa_msg)
        st.link_button("💬 Abrir WhatsApp", link, use_container_width=True)

        st.markdown("### ✅ Paso 3: Adjunta el Excel en WhatsApp")
        st.markdown(
            "- En el chat, toca el **clip 📎**\n"
            "- Elige **Documento**\n"
            "- Busca el Excel en **Archivos/Descargas** y envíalo."
        )


# ==========================
# STAR POINT
# ==========================
def reset_star_game_state():
    st.session_state.star_adv_count = 0
    st.session_state.star_golden_active = False
    st.session_state.golden_receiver = ""


def is_star_golden_now(modo_deuce: str) -> bool:
    return (
        modo_deuce == "Star Point"
        and (not st.session_state.in_tb)
        and st.session_state.pts[0] == 3
        and st.session_state.pts[1] == 3
        and st.session_state.adv == 0
        and st.session_state.star_golden_active
    )


# ==========================
# Orden sacadores por set (fuera TB)
# Reglas pedidas:
# - Game 1: eliges primer sacador (cualquiera)
# - Game 2: al terminar Game 1, eliges primer sacador SOLO del otro equipo
# - Game 3: automático el compañero del sacador del Game 1
# - Game 4: automático el compañero del sacador del Game 2
# ==========================
def reset_server_order_for_set():
    st.session_state.server_order = []
    st.session_state.server_index = 0
    st.session_state.team_first_server = {1: "", 2: ""}
    st.session_state.pending_other_team_pick = 0  # equipo que falta elegir (1 o 2)
    st.session_state.need_other_team_pick_now = False
    st.session_state.first_server_of_set = ""
    st.session_state.current_server = ""
    st.session_state.server_team = 0


def build_full_server_order(eq1: List[str], eq2: List[str]):
    t1 = st.session_state.team_first_server[1]
    t2 = st.session_state.team_first_server[2]
    if not t1 or not t2:
        return

    other1 = eq1[1] if eq1[0] == t1 else eq1[0]
    other2 = eq2[1] if eq2[0] == t2 else eq2[0]

    first_server = st.session_state.first_server_of_set
    first_team = equipo_de(first_server, eq1, eq2)

    if first_team == 1:
        order = [t1, t2, other1, other2]
    else:
        order = [t2, t1, other2, other1]

    st.session_state.server_order = order


def set_current_server_from_order(eq1: List[str], eq2: List[str]):
    if not st.session_state.server_order:
        return
    srv = st.session_state.server_order[st.session_state.server_index % 4]
    st.session_state.current_server = srv
    st.session_state.server_team = equipo_de(srv, eq1, eq2)


def advance_server_game(eq1: List[str], eq2: List[str]):
    # Si aún no hay orden completo y falta elegir el sacador del otro equipo,
    # NO cambiamos el sacador aquí. La UI lo pedirá cuando corresponda.
    if (not st.session_state.server_order) and st.session_state.pending_other_team_pick in (1, 2):
        return

    # Si ya hay orden completo, avanzamos al siguiente game
    if st.session_state.server_order:
        st.session_state.server_index = (st.session_state.server_index + 1) % 4
        set_current_server_from_order(eq1, eq2)


# ==========================
# TB: sacador automático (1,2,2,2...)
# ==========================
def tb_server_for_point(tb_point_index: int, rotation: List[str], start_idx: int) -> str:
    if not rotation:
        return ""
    if tb_point_index < 0:
        tb_point_index = 0

    if tb_point_index == 0:
        turn = 0
    else:
        turn = 1 + (tb_point_index - 1) // 2
    idx = (start_idx + turn) % len(rotation)
    return rotation[idx]


def ensure_tb_current_server(eq1: List[str], eq2: List[str]):
    if not st.session_state.in_tb:
        return

    tb_idx = st.session_state.tb_pts[0] + st.session_state.tb_pts[1]
    rotation = st.session_state.get("tb_rotation", [])
    start_idx = int(st.session_state.get("tb_start_idx", 0))

    if not rotation:
        st.session_state.current_server = ""
        st.session_state.server_team = 0
        return

    if len(rotation) < 4:
        srv = rotation[0]
        st.session_state.current_server = srv
        st.session_state.server_team = equipo_de(srv, eq1, eq2)
        return

    srv = tb_server_for_point(tb_idx, rotation, start_idx)
    st.session_state.current_server = srv
    st.session_state.server_team = equipo_de(srv, eq1, eq2)


# ==========================
# Match / marcador / TB
# ==========================
def reset_match():
    st.session_state.match_over = False

    st.session_state.sets = [0, 0]
    st.session_state.games = [0, 0]
    st.session_state.pts = [0, 0]
    st.session_state.adv = 0

    st.session_state.in_tb = False
    st.session_state.tb_tipo = ""
    st.session_state.tb_target = 0
    st.session_state.tb_pts = [0, 0]

    st.session_state.tb_rotation = []
    st.session_state.tb_start_idx = 0
    st.session_state.super_tb_first = ""
    st.session_state.super_tb_second = ""
    st.session_state.super_tb_ready = False

    reset_star_game_state()
    reset_server_order_for_set()


def set_actual() -> int:
    return max(sum(st.session_state.sets) + 1, 1)


def puntos_texto(eq_idx: int, modo_deuce: str) -> str:
    if st.session_state.in_tb:
        return str(st.session_state.tb_pts[eq_idx])

    p = st.session_state.pts[eq_idx]
    o = st.session_state.pts[1 - eq_idx]

    if modo_deuce in ("Advantage", "Star Point") and p == 3 and o == 3:
        return "AD" if st.session_state.adv == (eq_idx + 1) else "40"

    return PTS_TEXT.get(p, str(p))


def _activar_tb(tipo: str, target: int, eq1: List[str], eq2: List[str]):
    st.session_state.in_tb = True
    st.session_state.tb_tipo = tipo
    st.session_state.tb_target = target
    st.session_state.tb_pts = [0, 0]
    reset_star_game_state()

    if tipo == "SET":
        if st.session_state.server_order:
            next_idx = (st.session_state.server_index + 1) % 4
            st.session_state.tb_rotation = list(st.session_state.server_order)
            st.session_state.tb_start_idx = next_idx
        else:
            st.session_state.tb_rotation = []
            st.session_state.tb_start_idx = 0
        st.session_state.super_tb_ready = True

    elif tipo == "SUPER":
        st.session_state.tb_rotation = []
        st.session_state.tb_start_idx = 0
        st.session_state.super_tb_first = ""
        st.session_state.super_tb_second = ""
        st.session_state.super_tb_ready = False

    st.session_state.current_server = ""
    st.session_state.server_team = 0


def _terminar_tb_y_aplicar_ganador(eq_gana_tb: int):
    if st.session_state.tb_tipo == "SET":
        st.session_state.sets[eq_gana_tb - 1] += 1
    elif st.session_state.tb_tipo == "SUPER":
        st.session_state.sets[eq_gana_tb - 1] += 1
        st.session_state.match_over = True

    st.session_state.games = [0, 0]
    st.session_state.pts = [0, 0]
    st.session_state.adv = 0

    st.session_state.in_tb = False
    st.session_state.tb_tipo = ""
    st.session_state.tb_target = 0
    st.session_state.tb_pts = [0, 0]

    st.session_state.tb_rotation = []
    st.session_state.tb_start_idx = 0
    st.session_state.super_tb_first = ""
    st.session_state.super_tb_second = ""
    st.session_state.super_tb_ready = False

    reset_star_game_state()
    reset_server_order_for_set()


def _ganar_set_normal(eq_gana_set: int):
    st.session_state.sets[eq_gana_set - 1] += 1

    st.session_state.games = [0, 0]
    st.session_state.pts = [0, 0]
    st.session_state.adv = 0

    reset_star_game_state()
    reset_server_order_for_set()

    if st.session_state.formato_partido == "3 sets" and max(st.session_state.sets) >= 2:
        st.session_state.match_over = True
    if st.session_state.formato_partido == "Super tie-break" and max(st.session_state.sets) >= 2:
        st.session_state.match_over = True


def _check_activar_super_tb_si_corresponde(eq1: List[str], eq2: List[str]):
    if st.session_state.formato_partido != "Super tie-break":
        return
    if st.session_state.match_over:
        return
    if st.session_state.sets == [1, 1]:
        _activar_tb(tipo="SUPER", target=11, eq1=eq1, eq2=eq2)


def ganar_juego(eq_gana_game: int, eq1: List[str], eq2: List[str]):
    st.session_state.games[eq_gana_game - 1] += 1

    st.session_state.pts = [0, 0]
    st.session_state.adv = 0
    reset_star_game_state()

    g1, g2 = st.session_state.games

    if g1 == 6 and g2 == 6:
        _activar_tb(tipo="SET", target=7, eq1=eq1, eq2=eq2)
        return

    if (g1 >= 6 or g2 >= 6) and abs(g1 - g2) >= 2:
        _ganar_set_normal(1 if g1 > g2 else 2)
        _check_activar_super_tb_si_corresponde(eq1, eq2)
        return

    # ✅ FIX: Si se cerró un game y aún falta elegir sacador del otro equipo (Game 2),
    #        limpiamos el sacador para impedir guardar puntos hasta elegirlo.
    if (not st.session_state.server_order) and st.session_state.pending_other_team_pick in (1, 2):
        st.session_state.need_other_team_pick_now = True
        st.session_state.current_server = ""
        st.session_state.server_team = 0
        return

    # Avanza sacador SOLO si ya hay orden completo (server_order armado)
    advance_server_game(eq1, eq2)


def actualizar_marcador(eq_gana_punto: int, modo_deuce: str, eq1: List[str], eq2: List[str]):
    if st.session_state.match_over:
        return

    i = eq_gana_punto - 1
    j = 1 - i

    if st.session_state.in_tb:
        st.session_state.tb_pts[i] += 1
        a, b = st.session_state.tb_pts
        tgt = st.session_state.tb_target
        if (a >= tgt or b >= tgt) and abs(a - b) >= 2:
            _terminar_tb_y_aplicar_ganador(eq_gana_tb=eq_gana_punto)
        return

    if modo_deuce == "Golden":
        if st.session_state.pts[i] == 3 and st.session_state.pts[j] == 3:
            ganar_juego(eq_gana_punto, eq1, eq2)
            return
        st.session_state.pts[i] += 1
        if st.session_state.pts[i] >= 4 and (st.session_state.pts[i] - st.session_state.pts[j]) >= 2:
            ganar_juego(eq_gana_punto, eq1, eq2)
        return

    if modo_deuce == "Star Point":
        if is_star_golden_now(modo_deuce):
            ganar_juego(eq_gana_punto, eq1, eq2)
            return

        if st.session_state.pts[i] == 3 and st.session_state.pts[j] == 3 and st.session_state.adv == 0:
            st.session_state.adv = eq_gana_punto
            st.session_state.star_adv_count += 1
            if st.session_state.star_adv_count >= 2:
                st.session_state.star_golden_active = True
            return

        if st.session_state.pts[i] == 3 and st.session_state.pts[j] == 3 and st.session_state.adv != 0:
            if st.session_state.adv == eq_gana_punto:
                st.session_state.adv = 0
                ganar_juego(eq_gana_punto, eq1, eq2)
                return
            st.session_state.adv = 0
            return

        st.session_state.pts[i] += 1
        if st.session_state.pts[i] >= 4 and (st.session_state.pts[i] - st.session_state.pts[j]) >= 2:
            ganar_juego(eq_gana_punto, eq1, eq2)
        return

    if st.session_state.pts[i] == 3 and st.session_state.pts[j] == 3:
        if st.session_state.adv == 0:
            st.session_state.adv = eq_gana_punto
        elif st.session_state.adv == eq_gana_punto:
            st.session_state.adv = 0
            ganar_juego(eq_gana_punto, eq1, eq2)
        else:
            st.session_state.adv = 0
        return

    st.session_state.pts[i] += 1
    if st.session_state.pts[i] >= 4 and (st.session_state.pts[i] - st.session_state.pts[j]) >= 2:
        ganar_juego(eq_gana_punto, eq1, eq2)


# ==========================
# Excel (robusto + cronológico)
# ==========================
def get_excel_file() -> str:
    return st.session_state.excel_file


def _is_valid_xlsx(path: str) -> bool:
    if not os.path.exists(path):
        return False
    try:
        with zipfile.ZipFile(path, "r") as z:
            return "[Content_Types].xml" in z.namelist()
    except Exception:
        return False


def leer_eventos() -> pd.DataFrame:
    excel_file = get_excel_file()
    if not os.path.exists(excel_file) or not _is_valid_xlsx(excel_file):
        return pd.DataFrame(columns=EVENT_COLS)
    try:
        df = pd.read_excel(excel_file, sheet_name=SHEET_EVENTS, engine="openpyxl")
        for c in EVENT_COLS:
            if c not in df.columns:
                df[c] = ""
        return df[EVENT_COLS]
    except Exception:
        return pd.DataFrame(columns=EVENT_COLS)


def _aplicar_formato_fecha(excel_file: str):
    if not _is_valid_xlsx(excel_file):
        return
    wb = load_workbook(excel_file)
    for sheet in [SHEET_EVENTS, SHEET_SUMMARY]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        if "FechaHora" not in headers:
            continue
        col_idx = headers.index("FechaHora") + 1
        col_letter = get_column_letter(col_idx)
        for r in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{r}"]
            if cell.value:
                cell.number_format = "dd/mm/yyyy hh:mm:ss"
    wb.save(excel_file)


def guardar_excel(eventos_df: pd.DataFrame):
    excel_file = get_excel_file()

    if os.path.exists(excel_file) and not _is_valid_xlsx(excel_file):
        try:
            os.remove(excel_file)
        except Exception:
            pass

    tmp_dir = os.path.dirname(os.path.abspath(excel_file)) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=tmp_dir)
    os.close(fd)

    try:
        with pd.ExcelWriter(tmp_path, engine="openpyxl", mode="w") as writer:
            eventos_df.to_excel(writer, index=False, sheet_name=SHEET_EVENTS)
        _aplicar_formato_fecha(tmp_path)
        os.replace(tmp_path, excel_file)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


def insertar_evento_abajo(row: dict):
    df_old = leer_eventos()
    df_new = pd.DataFrame([row], columns=EVENT_COLS)
    df_out = pd.concat([df_old, df_new], ignore_index=True)
    guardar_excel(df_out)


def generar_resumen(eventos: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if eventos.empty:
        return pd.DataFrame(), pd.DataFrame()

    df = eventos.copy()
    for c in ["Saca", "SaqueEstado", "Resultado", "JugadorActor", "JugadorProvocador", "Asistencia", "Asistente"]:
        df[c] = df[c].fillna("").astype(str)

    df["FS_IN"] = (df["SaqueEstado"] == "Correcto").astype(int)
    df["FS_OUT"] = (df["SaqueEstado"].isin(["Primer error", "Doble falta"])).astype(int)
    df["SS_IN"] = (df["SaqueEstado"] == "Primer error").astype(int)
    df["SS_OUT"] = (df["SaqueEstado"] == "Doble falta").astype(int)
    df["DF"] = (df["SaqueEstado"] == "Doble falta").astype(int)

    df["WINNER"] = (df["Resultado"] == "Winner").astype(int)
    df["ENF"] = (df["Resultado"] == "Error no forzado").astype(int)
    df["EF"] = (df["Resultado"] == "Error forzado").astype(int)

    jugadores = sorted({
        x for x in (set(df["Saca"]) | set(df["JugadorActor"]) | set(df["JugadorProvocador"]) | set(df["Asistente"]))
        if x.strip()
    })

    def resumen_grupo(group_cols: List[str]) -> pd.DataFrame:
        rows = []
        grouped = df.groupby(group_cols, dropna=False) if group_cols else [((), df)]
        for keys, g in grouped:
            if not isinstance(keys, tuple):
                keys = (keys,)
            key_dict = dict(zip(group_cols, keys)) if group_cols else {}
            for jugador in jugadores:
                sub_srv = g[g["Saca"] == jugador]
                rows.append({
                    **key_dict,
                    "Jugador": jugador,
                    "1S_IN": int(sub_srv["FS_IN"].sum()),
                    "1S_OUT": int(sub_srv["FS_OUT"].sum()),
                    "2S_IN": int(sub_srv["SS_IN"].sum()),
                    "2S_OUT": int(sub_srv["SS_OUT"].sum()),
                    "DobleFalta": int(sub_srv["DF"].sum()),
                    "Winners": int(g[(g["Resultado"] == "Winner") & (g["JugadorActor"] == jugador)]["WINNER"].sum()),
                    "ENF": int(g[(g["Resultado"] == "Error no forzado") & (g["JugadorActor"] == jugador)]["ENF"].sum()),
                    "EF_Provocados": int(g[(g["Resultado"] == "Error forzado") & (g["JugadorProvocador"] == jugador)]["EF"].sum()),
                    "EF_Cometidos": int(g[(g["Resultado"] == "Error forzado") & (g["JugadorActor"] == jugador)]["EF"].sum()),
                    "Asistencias": int(g[(g["Asistencia"] == "Sí") & (g["Asistente"] == jugador)].shape[0]),
                })
        return pd.DataFrame(rows)

    resumen_set = resumen_grupo(["Set"]).sort_values(["Set", "Jugador"]).reset_index(drop=True)
    resumen_total = resumen_grupo([]).sort_values(["Jugador"]).reset_index(drop=True)
    return resumen_set, resumen_total


def guardar_resumen(eventos: pd.DataFrame, resumen_set: pd.DataFrame, resumen_total: pd.DataFrame):
    excel_file = get_excel_file()

    tmp_dir = os.path.dirname(os.path.abspath(excel_file)) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=tmp_dir)
    os.close(fd)

    try:
        with pd.ExcelWriter(tmp_path, engine="openpyxl", mode="w") as writer:
            eventos.to_excel(writer, index=False, sheet_name=SHEET_EVENTS)
            resumen_set.to_excel(writer, index=False, sheet_name=SHEET_SUMMARY, startrow=0)
            start = len(resumen_set) + 3
            resumen_total.to_excel(writer, index=False, sheet_name=SHEET_SUMMARY, startrow=start)

        _aplicar_formato_fecha(tmp_path)
        os.replace(tmp_path, excel_file)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


# ==========================
# Punto: estado + validación
# ==========================
def reset_punto():
    st.session_state.sel_saque_estado = ""
    st.session_state.sel_resultado = ""
    st.session_state.sel_golpe = ""
    st.session_state.sel_actor = ""
    st.session_state.sel_provocador = ""
    st.session_state.sel_asistencia = ""
    st.session_state.sel_asistente = ""
    st.session_state.point_start = None
    st.session_state.golden_receiver = ""


def start_timer_if_needed():
    if st.session_state.point_start is None:
        st.session_state.point_start = datetime.now()


def validar_punto(eq1: List[str], eq2: List[str], modo_deuce: str) -> Optional[str]:
    jugadores_set = set(eq1 + eq2)

    if not st.session_state.current_server:
        return "Falta sacador. Elígelo en el marcador."

    if is_star_golden_now(modo_deuce) and not st.session_state.golden_receiver:
        return "Falta seleccionar receptor para el Golden Point (Star Point)."

    # ✅ (3) SaqueEstado opcional: si viene, validar; si no, se asume Correcto al guardar
    se = st.session_state.sel_saque_estado
    if se and se not in SAQUE_ESTADOS:
        return "Estado del saque inválido."

    if se == "Doble falta":
        return None

    res = st.session_state.sel_resultado
    if res not in RESULTADOS:
        return "Falta seleccionar el resultado."

    golpe = st.session_state.sel_golpe
    if not golpe:
        return "Falta seleccionar el golpe."

    # Winner + Saque => actor auto
    if res == "Winner" and golpe == "Saque":
        return None

    # ✅ (2) ENF + Saque => actor auto (sacador)
    if res == "Error no forzado" and golpe == "Saque":
        return None

    if res == "Winner":
        if st.session_state.sel_actor not in jugadores_set:
            return "Falta seleccionar Actor."
        if st.session_state.sel_asistencia not in ("Sí", "No"):
            return "En Winner, indica si hubo asistencia (Sí/No)."
        if st.session_state.sel_asistencia == "Sí":
            asist = st.session_state.sel_asistente
            if asist not in jugadores_set:
                return "Asistente inválido."
            if asist == st.session_state.sel_actor:
                return "Asistente no puede ser igual al Actor."
            if equipo_de(asist, eq1, eq2) != equipo_de(st.session_state.sel_actor, eq1, eq2):
                return "Asistente debe ser del mismo equipo del Actor."
        return None

    if res == "Error no forzado":
        if st.session_state.sel_actor not in jugadores_set:
            return "Falta seleccionar Actor."
        return None

    if res == "Error forzado":
        actor = st.session_state.sel_actor
        prov = st.session_state.sel_provocador
        if actor not in jugadores_set:
            return "Falta seleccionar Actor (quién se equivocó)."
        if prov not in jugadores_set:
            return "Falta seleccionar Provocador (quién forzó)."
        if equipo_de(prov, eq1, eq2) == equipo_de(actor, eq1, eq2):
            return "Provocador debe ser del equipo contrario al Actor."
        return None

    return None


def make_row_base(modo_deuce: str) -> dict:
    # ✅ (4) Sin TB_Eq1 / TB_Eq2
    return {
        "FechaHora": datetime.now(),
        "Set": set_actual(),
        "Games_Eq1": st.session_state.games[0],
        "Games_Eq2": st.session_state.games[1],
        "Pts_Eq1": puntos_texto(0, modo_deuce),
        "Pts_Eq2": puntos_texto(1, modo_deuce),
        "TieBreak": bool(st.session_state.in_tb),
        "TB_Tipo": st.session_state.tb_tipo if st.session_state.in_tb else "",
    }


def registrar_evento(eq1: List[str], eq2: List[str], modo_deuce: str):
    if st.session_state.match_over:
        st.warning("El partido ya terminó. Reinicia si quieres registrar otro.")
        return

    err = validar_punto(eq1, eq2, modo_deuce)
    if err:
        st.error(err)
        return

    saca = st.session_state.current_server

    # ✅ (3) Si no se selecciona, asumimos "Correcto"
    saque_estado = st.session_state.sel_saque_estado or "Correcto"

    res = st.session_state.sel_resultado
    golpe = st.session_state.sel_golpe

    actor = st.session_state.sel_actor
    prov = st.session_state.sel_provocador
    asis = st.session_state.sel_asistencia
    asistente = st.session_state.sel_asistente

    if saque_estado == "Doble falta":
        eq_ganador = opuesto(equipo_de(saca, eq1, eq2))
        resultado_final = "Error no forzado"
        golpe_final = "Saque"
        actor_final = saca
        prov_final = ""
        asis_final = ""
        asistente_final = ""
    else:
        # ✅ (2) Winner+Saque y ENF+Saque => Actor auto sacador
        if (res == "Winner" and golpe == "Saque") or (res == "Error no forzado" and golpe == "Saque"):
            actor_final = saca
            prov_final = ""
            asis_final = ""
            asistente_final = ""
        else:
            actor_final = actor
            prov_final = prov if res == "Error forzado" else ""
            asis_final = asis if res == "Winner" else ""
            asistente_final = asistente if (res == "Winner" and asis == "Sí") else ""

        eq_ganador = ganador_equipo_por_regla(res, actor_final, prov_final, eq1, eq2)
        if eq_ganador not in (1, 2):
            st.error("No pude determinar el equipo ganador.")
            return

        resultado_final = res
        golpe_final = golpe

    actualizar_marcador(eq_ganador, modo_deuce, eq1, eq2)

    dur = mmss_from_start(st.session_state.point_start)

    row_point = {
        **make_row_base(modo_deuce),
        "Saca": saca,
        "SaqueEstado": saque_estado,
        "Resultado": resultado_final,
        "Golpe": golpe_final,
        "JugadorActor": actor_final,
        "JugadorProvocador": prov_final,
        "Asistencia": asis_final,
        "Asistente": asistente_final,
        "EquipoGanadorPunto": eq_ganador,
        "DuracionPunto": dur,
        "GoldenReceiver": st.session_state.golden_receiver if is_star_golden_now(modo_deuce) else "",
    }

    insertar_evento_abajo(row_point)
    reset_punto()
    st.success(f"Punto registrado. Ganó Equipo {eq_ganador} (auto).")


# ==========================
# UI
# ==========================
st.set_page_config(page_title="Padel Stats", layout="wide")
inject_css()
st.title("📊 Padel Stats — Registro por punto (local)")

# Sidebar: nombre archivo
st.sidebar.header("💾 Nombre del archivo")
fecha_archivo = st.sidebar.date_input("Fecha", value=date.today(), key="file_fecha")
jugador_archivo = st.sidebar.text_input("Jugador (referencia)", value="", key="file_jugador").strip()
campeonato = st.sidebar.text_input("Campeonato", value="", key="file_campeonato").strip()
categoria = st.sidebar.text_input("Categoría", value="", key="file_categoria").strip()

st.sidebar.caption(f"Fecha (visual): {fecha_archivo.strftime('%d/%m/%Y')}")
fecha_file = fecha_archivo.strftime("%d-%m-%Y")
parts = [fecha_file, jugador_archivo or "Jugador", campeonato or "Campeonato", categoria or "Categoria"]
parts = [sanitize_filename(p) for p in parts]
st.session_state.excel_file = "_".join(parts) + ".xlsx"
st.sidebar.write(f"Se guardará como: `{st.session_state.excel_file}`")
st.sidebar.divider()

# Sidebar: configuración
st.sidebar.header("⚙️ Configuración")
p1 = st.sidebar.text_input("Equipo 1 - Jugador A", value="Jugador1", key="p1")
p2 = st.sidebar.text_input("Equipo 1 - Jugador B", value="Jugador2", key="p2")
p3 = st.sidebar.text_input("Equipo 2 - Jugador A", value="Jugador3", key="p3")
p4 = st.sidebar.text_input("Equipo 2 - Jugador B", value="Jugador4", key="p4")

eq1 = [p1.strip(), p2.strip()]
eq2 = [p3.strip(), p4.strip()]
jugadores = unique_players(eq1, eq2)

modo_deuce_ui = st.sidebar.selectbox(
    "Modo en 40-40",
    ["—", "Advantage", "Golden", "Star Point"],
    index=0,
    key="modo_deuce_ui",
)
formato_ui = st.sidebar.selectbox(
    "Formato de partido",
    ["—", "3 sets", "Super tie-break"],
    index=0,
    key="formato_partido_ui",
)
st.session_state.modo_deuce = "" if modo_deuce_ui == "—" else modo_deuce_ui
st.session_state.formato_partido = "" if formato_ui == "—" else formato_ui
modo_deuce = st.session_state.get("modo_deuce", "")

# Init
if "initialized" not in st.session_state:
    st.session_state.initialized = True
    st.session_state.show_share = False
    reset_match()
    reset_punto()

# Gate
missing = []
if not st.session_state.get("modo_deuce"):
    missing.append("Modo en 40-40")
if not st.session_state.get("formato_partido"):
    missing.append("Formato de partido")
if missing:
    st.warning("Antes de comenzar el partido debes completar: " + ", ".join(missing))
    st.stop()

# Controles
st.sidebar.divider()
cA, cB = st.sidebar.columns(2)
if cA.button("🧹 Reset punto", use_container_width=True, key="btn_reset_punto"):
    reset_punto()
if cB.button("🔁 Reset partido", use_container_width=True, key="btn_reset_match"):
    reset_match()
    reset_punto()
    st.session_state.show_share = False

# Descargar excel
st.sidebar.divider()
excel_file = get_excel_file()
if os.path.exists(excel_file) and _is_valid_xlsx(excel_file):
    with open(excel_file, "rb") as f:
        st.sidebar.download_button("Descargar Excel", f, file_name=excel_file, use_container_width=True, key="dl_excel_sidebar")
else:
    st.sidebar.info("Todavía no hay archivo (registra al menos 1 punto).")

# Resumen
st.sidebar.divider()
if st.sidebar.button("🏁 Finalizar partido: generar resumen", use_container_width=True, key="btn_fin_resumen"):
    eventos = leer_eventos()
    rset, rtot = generar_resumen(eventos)
    if rset.empty and rtot.empty:
        st.sidebar.warning("No hay datos para resumir.")
        st.session_state.show_share = False
    else:
        guardar_resumen(eventos, rset, rtot)
        st.sidebar.success("Resumen guardado en hoja 'Resumen'.")
        st.session_state.show_share = True

# ==========================
# Marcador (con sacador visible y selector SOLO si aún no está elegido)
# IMPORTANTE: aquí NO se auto-selecciona el primero (usa allow_clear=True).
# ==========================
top1, top2 = st.columns([2, 3])

# Si está en TB, aseguramos sacador TB antes de mostrar
if st.session_state.get("in_tb", False):
    ensure_tb_current_server(eq1, eq2)

with top1:
    st.subheader("🏁 Marcador")

    sacador_actual = st.session_state.get("current_server", "")

    st.write(
        f"**Sets:** Eq1 {st.session_state.sets[0]} — Eq2 {st.session_state.sets[1]}\n\n"
        f"**Games:** Eq1 {st.session_state.games[0]} — Eq2 {st.session_state.games[1]}\n\n"
        f"**Puntos:** Eq1 {puntos_texto(0, modo_deuce)} — Eq2 {puntos_texto(1, modo_deuce)}\n\n"
        f"**Sacador:** **{sacador_actual or '—'}**"
    )

    # Solo aparece selector si NO hay sacador y NO estamos en TB
    if (not st.session_state.in_tb) and (not sacador_actual):
        st.warning("Elige el sacador del **Game 1** (quedará fijo durante el game).")
        elegido, _ = segmented_toggle(
            st,
            "Elegir sacador (Game 1)",
            state_key="pick_server_game1",
            options=jugadores,
            key="seg_sacador_game1_marcador",
            allow_clear=True,     # <- CRÍTICO: NO auto-selecciona el primero
        )
        if elegido:
            st.session_state.current_server = elegido
            st.session_state.server_team = equipo_de(elegido, eq1, eq2)

            # Configurar lógica de set (Game 1)
            st.session_state.first_server_of_set = elegido
            t = equipo_de(elegido, eq1, eq2)
            st.session_state.team_first_server[t] = elegido
            st.session_state.pending_other_team_pick = opuesto(t)
            st.session_state.need_other_team_pick_now = False
            st.session_state.server_index = 0

            # Limpiar el picker (para que no moleste luego)
            st.session_state.pick_server_game1 = ""
            reset_punto()
            st.rerun()

    if st.session_state.in_tb:
        st.info(
            f"**{('Super TB' if st.session_state.tb_tipo == 'SUPER' else 'Tie-break')}**: "
            f"Eq1 {st.session_state.tb_pts[0]} — Eq2 {st.session_state.tb_pts[1]}  "
            f"(a {st.session_state.tb_target}, dif 2)"
        )

    if st.session_state.match_over:
        st.success("✅ Partido terminado.")

with top2:
    st.subheader("⏱️ Timer del punto")
    st.write(f"Duración (hasta ahora): **{mmss_from_start(st.session_state.point_start)}**")

st.divider()

# ✅ FIX: Si aún no hay sacador (fuera TB), solo paramos si NO estamos en modo "elegir sacador Game 2"
if (not st.session_state.in_tb) and (not st.session_state.get("current_server", "")):
    if not (st.session_state.get("need_other_team_pick_now", False) and st.session_state.get("pending_other_team_pick", 0) in (1, 2)):
        st.info("Selecciona el sacador en el marcador para continuar.")
        st.stop()

# ==========================
# Sacador (reglas por set)
# - Durante el game, el sacador queda fijo (solo cambia al cerrar game).
# - Al cerrar Game 1: pide sacador del otro equipo (Game 2).
# - En Game 3 y 4: auto por orden [t1, t2, other1, other2].
# ==========================
st.subheader("1) Sacador (por game; set auto; TB auto)")

if st.session_state.in_tb:
    # TB: automático con rotación
    tb_idx = st.session_state.tb_pts[0] + st.session_state.tb_pts[1]

    if st.session_state.tb_tipo == "SUPER" and tb_idx == 0 and not st.session_state.super_tb_first:
        st.warning("Super tie-break: elige el 1º sacador (solo para el primer punto).")
        segmented_toggle(
            st,
            "1º sacador",
            state_key="super_tb_first",
            options=jugadores,
            key="seg_super_tb_first_only",
            allow_clear=True,
        )
        if st.session_state.super_tb_first:
            st.session_state.tb_rotation = [st.session_state.super_tb_first]
            st.session_state.tb_start_idx = 0
        st.stop()

    if st.session_state.tb_tipo == "SUPER" and tb_idx >= 1 and st.session_state.super_tb_first and not st.session_state.super_tb_second:
        team_first = equipo_de(st.session_state.super_tb_first, eq1, eq2)
        other_team_players = eq2 if team_first == 1 else eq1
        st.warning("Super tie-break: ahora elige el 2º sacador (del otro equipo).")
        segmented_toggle(
            st,
            "2º sacador",
            state_key="super_tb_second",
            options=other_team_players,
            key="seg_super_tb_second_late",
            allow_clear=True,
        )
        if st.session_state.super_tb_second:
            first = st.session_state.super_tb_first
            second = st.session_state.super_tb_second
            third = companero(first, eq1, eq2)
            fourth = companero(second, eq1, eq2)
            st.session_state.tb_rotation = [first, second, third, fourth]
            st.session_state.tb_start_idx = 0
            st.session_state.super_tb_ready = True
        st.stop()

    ensure_tb_current_server(eq1, eq2)
    st.info(f"Sacador TB (auto): **{st.session_state.current_server or '—'}** | Orden: {st.session_state.tb_rotation}")

else:
    # Si ya hay orden completo, solo mostramos (y el cambio ocurre al cerrar game)
    if st.session_state.server_order:
        st.info(f"Sacador actual (fijo durante el game): **{st.session_state.current_server}** — Orden set: {st.session_state.server_order}")
    else:
        # Aún no hay orden (falta elegir el sacador del otro equipo)
        st.info(f"Sacador actual (fijo durante el game): **{st.session_state.current_server or '—'}**")

    # Si al cerrar Game 1 corresponde elegir sacador del otro equipo para Game 2
    if st.session_state.need_other_team_pick_now and st.session_state.pending_other_team_pick in (1, 2):
        pending_team = st.session_state.pending_other_team_pick
        team_players = eq1 if pending_team == 1 else eq2
        st.warning(f"Fin de Game 1: elige sacador del **Equipo {pending_team}** (Game 2).")

        elegido2, _ = segmented_toggle(
            st,
            f"Sacador Equipo {pending_team} (Game 2)",
            state_key="pick_server_game2",
            options=team_players[:2],
            key=f"seg_sacador_game2_{pending_team}",
            allow_clear=True,   # <- no auto-selecciona el primero
        )
        if elegido2:
            st.session_state.team_first_server[pending_team] = elegido2

            # Armar orden completo: [t1,t2,other1,other2] o [t2,t1,other2,other1]
            build_full_server_order(eq1, eq2)

            # Game 2 debe ser el segundo de la lista => index=1
            st.session_state.server_index = 1
            set_current_server_from_order(eq1, eq2)

            # Limpieza flags
            st.session_state.pending_other_team_pick = 0
            st.session_state.need_other_team_pick_now = False
            st.session_state.pick_server_game2 = ""

            reset_punto()
            st.rerun()

st.divider()

# Star Point receptor
if is_star_golden_now(modo_deuce):
    st.warning("⭐ Star Point: este punto es GOLDEN. Elige receptor.")
    receiving_team = opuesto(equipo_de(st.session_state.current_server, eq1, eq2))
    receivers = eq1 if receiving_team == 1 else eq2
    segmented_toggle(
        st,
        "Receptor (Golden)",
        state_key="golden_receiver",
        options=receivers[:2],
        key="seg_golden_receiver",
        allow_clear=True,
    )
    st.write(f"Receptor: **{st.session_state.golden_receiver or '—'}**")
    st.divider()

# ==========================
# Estado saque
# ✅ (1) UI solo con "Error 1er saque" (y dejamos doble falta para registrarla)
# ✅ (3) opcional: no se exige para guardar
# ==========================
st.subheader("1.1) Estado del saque (opcional)")
SAQUE_UI = ["❌ Error 1er saque", "❌❌ Doble falta"]
MAP_UI_TO_VAL = {
    "❌ Error 1er saque": "Primer error",
    "❌❌ Doble falta": "Doble falta",
    "": "",
}
MAP_VAL_TO_UI = {v: k for k, v in MAP_UI_TO_VAL.items() if v}
st.session_state._tmp_saque_ui = MAP_VAL_TO_UI.get(st.session_state.sel_saque_estado, "")

new_ui, _ = segmented_toggle(
    st,
    "Estado del saque",
    state_key="_tmp_saque_ui",
    options=SAQUE_UI,
    key="seg_saque_estado",
    allow_clear=True,
)
st.session_state.sel_saque_estado = MAP_UI_TO_VAL.get(new_ui, "")

if st.session_state.sel_saque_estado:
    start_timer_if_needed()

if st.session_state.sel_saque_estado == "Doble falta":
    st.session_state.sel_resultado = ""
    st.session_state.sel_golpe = ""
    st.session_state.sel_actor = ""
    st.session_state.sel_provocador = ""
    st.session_state.sel_asistencia = ""
    st.session_state.sel_asistente = ""

st.divider()

if st.session_state.sel_saque_estado == "Doble falta":
    st.subheader("✅ Guardar punto (Doble falta)")
    if st.button("Guardar doble falta (auto)", type="primary", use_container_width=True, key="guardar_df"):
        registrar_evento(eq1, eq2, modo_deuce)

else:
    st.subheader("2) Resultado del punto")
    new_res, old_res = segmented_toggle(
        st,
        "Resultado",
        state_key="sel_resultado",
        options=RESULTADOS,
        key="seg_resultado",
        allow_clear=True,
    )
    if new_res != old_res:
        st.session_state.sel_actor = ""
        st.session_state.sel_provocador = ""
        st.session_state.sel_asistencia = ""
        st.session_state.sel_asistente = ""
        if new_res == "":
            st.session_state.sel_golpe = ""

    st.subheader("3) Tipo de golpe")
    segmented_toggle(
        st,
        "Golpe",
        state_key="sel_golpe",
        options=GOLPES,
        key="seg_golpe",
        allow_clear=True,
    )

    # ✅ (2) Winner+Saque y ENF+Saque => Actor auto (sacador)
    if st.session_state.sel_golpe == "Saque" and st.session_state.sel_resultado in ("Winner", "Error no forzado"):
        st.info(
            f"{st.session_state.sel_resultado} de **saque**: Actor = sacador (**{st.session_state.current_server}**) (auto)."
        )
        st.session_state.sel_actor = ""
        st.session_state.sel_provocador = ""
        st.session_state.sel_asistencia = ""
        st.session_state.sel_asistente = ""

    st.divider()
    st.subheader("4) Jugadores involucrados (según resultado)")
    res = st.session_state.sel_resultado

    if res in ("Winner", "Error no forzado"):
        if st.session_state.sel_golpe == "Saque" and res in ("Winner", "Error no forzado"):
            st.write("Actor: **(auto) Sacador**")
        else:
            st.write("Elige **Actor**:")
            segmented_toggle(
                st,
                "Actor",
                state_key="sel_actor",
                options=jugadores,
                key=f"seg_actor_{res}",
                allow_clear=True,
            )

            if res == "Winner" and st.session_state.sel_actor:
                st.divider()
                st.subheader("5) Asistencia (solo Winner)")
                new_asis, old_asis = segmented_toggle(
                    st,
                    "Asistencia",
                    state_key="sel_asistencia",
                    options=["Sí", "No"],
                    key="seg_asistencia",
                    allow_clear=True,
                )
                if new_asis != old_asis:
                    if st.session_state.sel_asistencia == "Sí":
                        st.session_state.sel_asistente = companero(st.session_state.sel_actor, eq1, eq2)
                    else:
                        st.session_state.sel_asistente = ""
                if st.session_state.sel_asistencia == "Sí":
                    st.info(f"Asistente asumido automáticamente: **{st.session_state.sel_asistente}**")

    elif res == "Error forzado":
        st.write("Elige **Actor** (quién se equivocó):")
        segmented_toggle(
            st,
            "Actor (error)",
            state_key="sel_actor",
            options=jugadores,
            key="seg_actor_ef",
            allow_clear=True,
        )

        if st.session_state.sel_actor:
            actor_team = equipo_de(st.session_state.sel_actor, eq1, eq2)
            st.write("Ahora elige **Provocador** (quién forzó el error):")
            prov_options = [j for j in jugadores if equipo_de(j, eq1, eq2) != actor_team]
            segmented_toggle(
                st,
                "Provocador",
                state_key="sel_provocador",
                options=prov_options,
                key="seg_provocador",
                allow_clear=True,
            )

    st.divider()
    st.subheader("✅ Guardar punto")
    if st.button("Guardar punto (auto)", type="primary", use_container_width=True, key="guardar_punto"):
        registrar_evento(eq1, eq2, modo_deuce)

st.divider()

# Mostrar guía de compartir tras generar resumen
if st.session_state.get("show_share", False):
    ui_compartir_excel_con_guia(get_excel_file())
    st.divider()

st.subheader("📄 Últimos eventos guardados")
df_show = leer_eventos()
st.dataframe(df_show.tail(30).iloc[::-1], use_container_width=True)
st.caption("Se guarda en Excel al registrar cada punto. El resumen se genera en la sidebar.")
